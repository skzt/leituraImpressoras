import openpyxl as opxl
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from leiturasDigital import LeiturasDigital
from controlePoli import ControlePoli
from configs import Configs


class MainWindow(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        self.minsize(width=400, height=100)
        self.title("Leitura de Impressões")

        self.digitalVar = tk.StringVar()
        self.controleVar = tk.StringVar()
        self.configs = Configs()

        ##############################################################################
        # Seleção de planilha da Digital World
        ##############################################################################
        tk.Label(self, text="Selecione a Planilha da Digital").grid(row=0, column=0)
        tk.Entry(self, textvariable=self.digitalVar,
                 state='readonly',
                 width=70).grid(row=1, column=0, padx=(10, 5))

        selectDigitalButton = tk.Button(self)
        selectDigitalButton['text'] = "Abrir"
        selectDigitalButton['command'] = lambda: self.digitalVar.set(
            filedialog.askopenfilename(
                filetypes=(
                    ("Arquivo Excel", "*.xlsx"),
                    ("All files", "*.*"))))
        selectDigitalButton.grid(row=1, column=1, padx=(5, 10))

        ##############################################################################
        # Seleção de planilha da Controle Polipeças
        ##############################################################################
        tk.Label(self, text="Selecione a Planilha de Controle").grid(row=2, column=0)
        tk.Entry(self, textvariable=self.controleVar,
                 state='readonly',
                 width=70).grid(row=3, column=0, padx=(10, 5))

        selectControleButton = tk.Button(self)
        selectControleButton['text'] = "Abrir"
        selectControleButton['command'] = lambda: self.controleVar.set(filedialog.askopenfilename(
            filetypes=(
                ("Arquivo Excel", "*.xlsx"),
                ("All files", "*.*"))))
        selectControleButton.grid(row=3, column=1, padx=(5, 10))

        ##############################################################################

        executarButton = tk.Button(self)
        executarButton['text'] = "Executar"
        executarButton['command'] = self.executarLeitura
        executarButton.grid(row=4, column=0, padx=(10, 0), pady=(10, 10), sticky='w')

        sairButton = tk.Button(self)
        sairButton['text'] = "Sair"
        sairButton['command'] = self.destroy
        sairButton.grid(row=4, column=1, padx=(0, 10), pady=(10, 10))

    def executarLeitura(self):
        wbDigital = opxl.load_workbook(self.digitalVar.get(), data_only=True)

        sheetVariable = tk.StringVar()
        sheetVariable.set(wbDigital.sheetnames[0])

        if len(wbDigital.sheetnames) > 1:
            """
                Cria uma janela auxiliar, solicitando ao usuario que selecione
                a aba da planilha Excel em que estão os dados de impressão.
            """
            sheetWindow = tk.Toplevel(self)

            sheetLabel = tk.Label(sheetWindow)
            sheetLabel['text'] = "Selecione a aba correta:"
            sheetLabel.pack()

            sheetBox = ttk.Combobox(sheetWindow)
            sheetBox['values'] = wbDigital.sheetnames
            sheetBox['textvariable'] = sheetVariable
            sheetBox['state'] = 'readonly'
            sheetBox.bind('<<ComboboxSelected>>', lambda _: sheetWindow.destroy())
            sheetBox.pack()

            self.wait_window(sheetWindow)
        try:
            leitura = LeiturasDigital(self.configs, wbDigital[sheetVariable.get()])
        except AttributeError as err:
            messagebox.showerror(title="Erro ao Durante leitura Arquivo",
                                 message=err,
                                 parent=self)

            return

        wbControle = opxl.load_workbook(self.controleVar.get())
        escrita = ControlePoli(self.configs, wbControle)
        if escrita.prepararPlanilha():
            try:
                wbControle.save(self.controleVar.get())

            except OSError as err:
                msg1 = "Erro ao salvar o arquivo, planilha '%s' " \
                       "esta sendo usada por outro processo.\n" \
                       % str(err).split("'")[1]

                msg2 = "Verifique se a planilha esta fechada, " \
                       "finalize outros processos que possam estar utilizando " \
                       "esta planilha e tente novamente."

                msg = msg1 + msg2

                messagebox.showerror(title="Erro ao Salvar Arquivo",
                                     message=msg,
                                     parent=self)

                return
        # self.destroy()
        verificaEscrita = escrita.escreverLeituras(leitura)

        if verificaEscrita is True:
            wbControle.save(self.controleVar.get())
            messagebox.showinfo(title="Concluido",
                                message="Executado com Sucesso!",
                                parent=self)
            self.destroy()
        else:
            msg = "Número do serie (%s) de %s - %s não econtrado na planilha %s.\n Abortando!" % (verificaEscrita[2],
                                                                                                  verificaEscrita[1],
                                                                                                  verificaEscrita[0],
                                                                                                  self.controleVar.get()
                                                                                                  )
            messagebox.showerror(title="Erro Número de Série",
                                 message=msg,
                                 parent=self)
            self.destroy()


if __name__ == "__main__":
    mw = MainWindow()
    mw.mainloop()
